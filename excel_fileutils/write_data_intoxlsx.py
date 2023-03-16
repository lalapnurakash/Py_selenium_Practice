import openpyxl as op
import pandas as pd
from openpyxl.styles import PatternFill

df1=pd.read_excel('D:\\s3reports\\UAT.xlsx')

df2=pd.read_excel('D:\\s3reports\\prod.xlsx')

# filling  missing data with N/A
df=0
if len(df2)>len(df1):
    df = df2
# add missing rows to the existing dataframe with default value 'N/A'
    df1 = df1.reindex(range(len(df2)), fill_value='N/A')
else:
    df = df1
# add missing rows to the existing dataframe with default value 'N/A'
    df = df1.reindex(range(len(df1)), fill_value='N/A')

#create new workbook
wb=op.Workbook()
ws=wb.active
cols=df.columns.values
for row in ws.iter_rows(max_row=1, max_col=len(df.columns)):
 i=0
 for cell in row:
    cell.value = cols[i]
    cell.fill=PatternFill(bgColor="FFC7CE", fill_type = "solid")
    i=i+1

num_rows, num_cols = df.shape

for rw in range(1,num_rows+1):
    for cl in range(1,num_cols+1):
        # get the values of the two cells being compared
         value1 = df1.iloc[rw-1, cl-1]
         value2 = df2.iloc[rw-1, cl-1]
         if  rw!=0:
           if value1 == value2:
             (ws.cell(row=rw+1,column=cl)).value='True'
           else:
             (ws.cell(row=rw+1,column=cl)).value='False'
         if cl==1 and rw!=0:
             (ws.cell(row=rw+1, column=cl)).value=str(value1)+""+str(value2)
         if rw!=0 and cl==3:
             (ws.cell(row=rw+1, column=cl)).value = value1

wb.save('demo.xlsx')





