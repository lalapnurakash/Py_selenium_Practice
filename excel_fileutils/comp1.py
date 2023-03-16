from typing import Any

import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook

df1=pd.read_excel('D:\\s3reports\\UAT.xlsx')

df2=pd.read_excel('D:\\s3reports\\prod.xlsx')

# filling  missing data with N/A
df=0
if len(df2)>len(df1):
    df = df2
# add missing rows to the existing dataframe with default value 'N/A'
    df1 = df1.reindex(range(len(df2)), fill_value='N/A')
else:
    df = df1
# add missing rows to the existing dataframe with default value 'N/A'
    df = df1.reindex(range(len(df1)), fill_value='N/A')

num_rows, num_cols = df.shape
new_df = pd.DataFrame()

#new_df.set_axis(df.columns.to_list(),axis=1,inplace=True)
#new_df = new_df.append(df1)
#new_df = pd.concat([new_df,df1], axis=1)
#new_df=pd.DataFrame(columns=df.columns)

#new_df.columns=df.columns
# Load the Excel file
#new_df=df.copy()
# iterate over the rows and columns of the dataframes

# function definition

#new_df = new_df.reset_index(drop=False)
#new_df.set_index(df.columns)
for row in range(num_rows):
    for col in range(num_cols):
        # get the values of the two cells being compared

         wb = load_workbook(filename='example.xlsx')
         ws = wb.active

         value1 = df1.iloc[row, col]
         value2 = df2.iloc[row, col]
         if col!=0  :
           if value1 == value2:
             new_df.
             new_df.loc[row,col]='True'
             print(f'Cell ({row}, {col}) is equal')
           else:
            # ws.cell(row=row, column=col, value='False')
             print(f'Cell ({row}, {col}) is not equal')
             new_df.loc[row,col]='False'
         else:
             new_df.loc[row,col]=str(value1)+""+str(value2)

# copy data from 2nd comparison file
#new_df = pd.concat([new_df,df2], axis=1)
def apply_colour(value):
  if value =='False':
      colour = '#00B8EA'  # Blue


#new_df = new_df.reset_index(drop=True)
#new_df.columns = range(len(new_df.columns))
#new_df = new_df.style.applymap(apply_colour)
#new_df = new_df.append(df2)
#new_df = pd.DataFrame(columns=df.columns.to_list())
#new_df=new_df.rename((df.columns.to_list))

new_df.set_axis(df.columns,axis=1,inplace=True)
new_df.to_excel('new.xlsx',index=False)

