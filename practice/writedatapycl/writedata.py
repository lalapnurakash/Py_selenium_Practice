import openpyxl
from openpyxl.reader.excel import load_workbook

wb=openpyxl.Workbook()
ws=wb.active
wb.create_sheet("sample")
w=wb.remove(wb['sample'])

#ws['A4'] = 4
#ws.cell(row=4, column=1, value=10)
#print(ws.cell(row=4, column=1))
# sheet["A1"] = "hello"
# sheet["B1"] = "world!"


wb.save("newxl.xlsx")
print(wb.sheetnames)
# ws.title = "New Title"

'''to save xl as template'''

# wb = load_workbook('newxl.xlsx')
# wb.template = True
# wb.save('document_template.xltx')
# wb = load_workbook('document_template.xltx')
# wb.template = False
# wb.save('document.xlsx', as_template=False)