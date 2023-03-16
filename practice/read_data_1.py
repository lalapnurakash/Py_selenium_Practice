import openpyxl.workbook
from openpyxl import load_workbook

path="C:\\Users\\1003633\\Desktop\\New folder\\book2.xlsx"
wb = load_workbook(filename=path)
s1=wb.worksheets[0]
sheet=wb.active
print(wb.sheetnames)
print(sheet)

print(sheet.cell(row=1,column=1).value)
for row in sheet.iter_cols(min_row=2, max_row=2, min_col=0, max_col=3, values_only=True):
    print(row)


