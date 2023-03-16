import json

import self
from openpyxl.reader.excel import load_workbook
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import ExcelWriter

from sample1.readwithopenpyxl import row

sheet=""
def open_workbook(path):
    global sheet
    workbook = load_workbook(filename=path)
    print(f"Worksheet names: {workbook.sheetnames}")
    #sheet = workbook.active
    #sheet=workbook.worksheets[0]
   # sheet1 = workbook["Sheet1"]
    sheet=workbook.worksheets[0]
    # sheet["A1"] = "hello"
    # sheet["B1"] = "world!"
    print(sheet["A1"].value)
def readrows(self,sheet):
    for row in sheet.iter_rows(min_row=0,max_row=2,min_col=0,max_col=2,values_only=True):
        print(row.value)
def readcols(self,sheet):
    for row in sheet.iter_cols(min_row=0, max_row=2, min_col=0, max_col=2, values_only=True):
        print(row)
def converttojson(self,sheet):
    products={}
    for row in sheet.iter_rows(min_row=2, min_col=1,  values_only=True):
        name=row[0]
        product={
          "age":row[1],
          "gender":row[2],
          "id_val":row[3]
        }
        products[name]=product
    print(json.dumps(products))



if __name__ == "__main__":
    open_workbook("C:\\Users\\1003633\\Desktop\\New folder\\Book1.xlsx")
    converttojson(self,sheet)

